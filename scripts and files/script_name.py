import openpyxl
import psycopg2
from psycopg2 import sql
from tkinter import filedialog
from tkinter import Tk
from datetime import datetime
import getpass  # Import the getpass module

# PostgreSQL connection details
db_server = "enter your server" #e.g. yourdwh.yourdomain.com.au
db_port = 1234 # replace 1234 with the port number
db_name = "Enter database name" # exampleDWH
db_user = "enter username" #e.g. postgres1
db_password = getpass.getpass("Enter your database password: ")  # Prompt user for password

# Function to check table existence in a specific schema
def table_exists_in_schema(cursor, schema_name, table_name):
    query = sql.SQL("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_schema = %s AND table_name = %s)").format(
        sql.Identifier(schema_name), sql.Identifier(table_name))
    cursor.execute(query, (schema_name, table_name))
    return cursor.fetchone()[0]

# Function to check field existence in a table in a specific schema
def field_exists_in_schema(cursor, schema_name, table_name, field_name):
    query = sql.SQL("SELECT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema = %s AND table_name = %s AND column_name = %s)").format(
        sql.Identifier(schema_name), sql.Identifier(table_name), sql.Identifier(field_name))
    cursor.execute(query, (schema_name, table_name, field_name))
    return cursor.fetchone()[0]

# Custom function for case-insensitive schema name conversion
def convert_to_valid_schema(user_input):
    valid_schemas = ["HumRes", "Finance", "Test", "public"]
    lowercased_input = user_input.lower()
    
    for schema in valid_schemas:
        if lowercased_input == schema.lower():
            return schema

    print("Invalid schema name. Please enter a valid schema name.")
    exit()

# Connect to PostgreSQL
try:
    # Prompt user for the schema and convert to valid format
    user_schema = convert_to_valid_schema(input("Enter the table schema e.g Finance: "))


    # Open file dialog to select Excel file
    root = Tk()
    root.withdraw()  # Hide the main window
    xlsx_file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])

    connection = psycopg2.connect(
        host=db_server,
        port=db_port,
        database=db_name,
        user=db_user,
        password=db_password
    )

    with connection.cursor() as cursor:
        views = []
        missing_tables = set()

        # Set to store unique tables and fields encountered in the Excel file
        excel_tables_and_fields = set()

        # Read Excel file and check table and field existence
        workbook = openpyxl.load_workbook(xlsx_file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            table_name = None
            skip_table_description = False

            for row in sheet.iter_rows(min_row=2, values_only=True):
                cell_value = row[0].strip() if row[0] else None

                if cell_value and cell_value.startswith("v_"):
                    table_name = cell_value
                    skip_table_description = True
                    continue

                if table_name and cell_value == "Field":
                    skip_table_description = False
                    continue

                if skip_table_description:
                    continue

                if table_name and cell_value:
                    field_name = cell_value.strip()
                    if table_name.startswith("v_"):
                        category_list = views
                        excel_tables_and_fields.add((table_name, field_name))
                    else:
                        continue

                    table_existence = table_exists_in_schema(cursor, user_schema, table_name)
                    category_list.append((table_name, table_existence, field_name,
                                          field_exists_in_schema(cursor, user_schema, table_name, field_name) if field_name else None))

        # Check for missing tables and fields in the data warehouse
        cursor.execute(sql.SQL("SELECT table_name, column_name FROM information_schema.columns WHERE table_schema = %s"), (user_schema,))
        warehouse_tables_and_fields = set(cursor.fetchall())

        missing_tables_and_fields = warehouse_tables_and_fields - excel_tables_and_fields

        # Print results based on category
        print("\nResults:")
        print("\033[94mViews:\033[0m")

        # Print existing views in alphabetical order
        views.sort(key=lambda x: x[0])
        for table, exists, field, field_existence in views:
            print(f"{table}: {'\033[92mExists\033[0m' if exists else '\033[91mDoes not exist\033[0m'}")
            if field:
                print(f"  - {field}: {'\033[92mExists\033[0m' if field_existence else '\033[91mDoes not exist\033[0m'}")

        # Print missing tables and fields for views in alphabetical order
        missing_tables_and_fields = sorted(
            [item for item in missing_tables_and_fields if item[0].startswith("v_")],
            key=lambda x: (x[0], x[1])
        )
        print("\n\033[94mMissing Tables and Fields:\033[0m")

        current_table = None  # To keep track of the current table being printed
        for table_name, field_name in missing_tables_and_fields:
            if table_name != current_table:
                if current_table:
                    print()  # Add a newline between different tables
                current_table = table_name

            if field_name:
                print(f"{table_name} - {field_name}: \033[91mMissing in Excel file\033[0m")
            else:
                print(f"{table_name}: \033[91mMissing in Excel file\033[0m")

        # Create a new Excel workbook
        output_workbook = openpyxl.Workbook()

        # Remove the default "Sheet" created by openpyxl
        default_sheet = output_workbook["Sheet"]
        output_workbook.remove(default_sheet)

        # Create sheets in the workbook
        views_sheet = output_workbook.create_sheet("Data Dictionary")
        missing_tables_sheet = output_workbook.create_sheet("Data Warehouse")

        # Add header rows to the sheets
        views_sheet.append(["Table", "Table Existence in DWH", "Field", "Field Existence in DWH"])
        missing_tables_sheet.append(["Table", "Field", "Status"])

        # Print results based on category to the 'Views' sheet
        for table, exists, field, field_existence in views:
            views_sheet.append([table, 'Exists' if exists else 'Does not exist', field, 'Exists' if field_existence else 'Does not exist' if field else ''])

        # Print missing tables and fields for views to the 'Missing Tables and Fields' sheet
        for table_name, field_name in missing_tables_and_fields:
            missing_tables_sheet.append([table_name, field_name, 'Missing in Excel file' if field_name else 'Missing in Excel file'])

        # Prompt user to choose where to save the .xlsx file
        output_filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        # Save the output Excel file
        output_workbook.save(output_filename)
        print(f"\nResults have been saved to '{output_filename}'.")

except psycopg2.Error as e:
    print(f"\033[91mError: {e}\033[0m")

finally:
    if connection:
        connection.close()